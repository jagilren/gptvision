#24 Mayo 2025 Este archivo se debe eliminar es solo de pruebas
import openai
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

# Configura tu clave de API de OpenAI


# Ruta a la imagen que contiene la tabla
ruta_imagen = "WhatsApp Image 2025-05-23 at 06.30.43.jpeg"

# 1. Convertir imagen a texto usando OpenAI
def extraer_tabla_desde_imagen(imagen_path):
    with open(imagen_path, "rb") as img_file:
        response = openai.Image.create_variation(
            image=img_file,
            model="gpt-4o",  # asegúrate que tu cuenta soporte vision
            prompt=(
                "Extrae los datos debajo de los encabezados 'Date', 'Time', 'Value', 'Absorbance'. "
                "Agrega el año 2025 al campo de fecha. Devuelve los registros en formato CSV, sin encabezados."
            ),
            response_format="text"
        )
    return response["data"]

# 2. Guardar en DataFrame
def procesar_texto_a_df(texto_csv):
    from io import StringIO
    df = pd.read_csv(StringIO(texto_csv), header=None)
    df.columns = ["Date", "Time", "Value", "Absorbance"]
    return df

# 3. Agregar al archivo Excel
def agregar_a_excel(df, ruta_excel, hoja="WAD Entrada"):
    try:
        libro = load_workbook(ruta_excel)
        if hoja in libro.sheetnames:
            hoja_activa = libro[hoja]
        else:
            hoja_activa = libro.create_sheet(hoja)
    except FileNotFoundError:
        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=hoja, index=False)
        return

    # Encontrar la última fila
    ultima_fila = hoja_activa.max_row + 1
    for fila in df.itertuples(index=False):
        hoja_activa.append(list(fila))

    libro.save(ruta_excel)

# 4. Ejecutar todo
def main():
    texto_extraido = extraer_tabla_desde_imagen(ruta_imagen)
    df = procesar_texto_a_df(texto_extraido)
    agregar_a_excel(df, "WAD Entrada.xlsx")

# Llamar la función principal
main()
