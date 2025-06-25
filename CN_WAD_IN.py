#24 Mayo 2025 Este archivo se debe eliminar es solo de pruebas borrar
import openai
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from pathlib import Path

#1 Configura tu clave de API de OpenAI


# 2. Guardar en DataFrame
def procesar_texto_a_df(texto_csv):
    from io import StringIO
    df = pd.read_csv(StringIO(texto_csv), header=None)
    df.columns = ["Date", "Time", "Value", "Absorbance"]
    return df

# 3. Agregar al archivo Excel
def agregar_a_excel(df, ruta_excel, hoja="Datos_WAD"):
    try:
        libro = load_workbook(ruta_excel)
        if hoja in libro.sheetnames:
            hoja_activa = libro[hoja]
        else:
            hoja_activa = libro.create_sheet(hoja)
    except FileNotFoundError:
        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=hoja, index=False)
        return

    # Encontrar la última fila
    ultima_fila = hoja_activa.max_row + 1
    for fila in df.itertuples(index=False):
        hoja_activa.append(list(fila))

    libro.save(ruta_excel)

# 4. Ejecutar todo
def main(respuesta,RUTA_EXCEL):

    df = procesar_texto_a_df(respuesta)
    agregar_a_excel(df, RUTA_EXCEL)


# Llamar la función principal

